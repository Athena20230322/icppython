import pandas as pd
import os
from openpyxl import load_workbook
import shutil
import warnings
import re  # 引入正則表達式，用來處理多種分隔符號

# 1. 忽略警告
warnings.simplefilter(action='ignore', category=FutureWarning)

# 2. 設定路徑
source_folder = r"C:\Users\p10381190\Desktop\種子測試表單\0109-0113測試同仁回饋"
output_file = r"C:\Users\p10381190\Desktop\種子測試表單\測試回饋合併.xlsx"

all_data_list = []
total_staff_set = set()  # 【修改】使用集合 (Set) 來儲存唯一不重複的人名
template_file = None

print("正在彙整測試回饋（保留名單並合併資訊）...")

if not os.path.exists(source_folder):
    print(f"❌ 找不到路徑: {source_folder}")
else:
    for filename in os.listdir(source_folder):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(source_folder, filename)
            if template_file is None: template_file = file_path

            try:
                # --- 【核心修改：拆解多個名字】 ---
                # 1. 移除前綴與副檔名
                clean_name = filename.replace('.xlsx', '').replace('測試回饋-', '').replace('測試回饋_', '')

                # 2. 使用正則表達式拆解名字 (支援 、 , 空格 等分隔符)
                # re.split 會根據 [ ] 內的字元進行切割
                individual_names = re.split(r'[、,，\s]+', clean_name)

                # 3. 移除空字串並加入總集合
                for name in individual_names:
                    if name.strip():
                        total_staff_set.add(name.strip())

                # 讀取數據 (A5為標題，所以 header=4)
                df = pd.read_excel(file_path, header=4)
                df = df.dropna(subset=[df.columns[0]])
                all_data_list.append(df)

                print(f"✅ 已讀取: {filename}")
            except Exception as e:
                print(f"❌ 讀取 {filename} 出錯: {e}")

    if all_data_list:
        # 3. 合併數據 (與您原本邏輯一致)
        combined_df = pd.concat(all_data_list, ignore_index=True)
        combined_df['has_data'] = combined_df.iloc[:, 1:].notna().any(axis=1)

        df_with_info = combined_df[combined_df['has_data'] == True]
        df_no_info = combined_df[combined_df['has_data'] == False]
        testers_with_info = set(df_with_info.iloc[:, 0].unique())

        df_blank_to_keep = df_no_info[~df_no_info.iloc[:, 0].isin(testers_with_info)].drop_duplicates(
            subset=[df_no_info.columns[0]])

        final_df = pd.concat([df_with_info, df_blank_to_keep], ignore_index=True)
        final_df = final_df.drop(columns=['has_data'])
        final_df = final_df.sort_values(by=final_df.columns[0])

        # 4. 寫入 Excel
        shutil.copyfile(template_file, output_file)
        book = load_workbook(output_file)
        sheet = book.active

        if sheet.max_row > 5:
            sheet.delete_rows(6, sheet.max_row)

        for r_idx, row in enumerate(final_df.values, start=6):
            for c_idx, value in enumerate(row, start=1):
                val = "" if pd.isna(value) else value
                sheet.cell(row=r_idx, column=c_idx, value=val)

        try:
            book.save(output_file)

            # 將集合轉換為排序後的列表，方便列印
            sorted_staff_list = sorted(list(total_staff_set))

            print("\n" + "=" * 50)
            print("✨ 任務圓滿完成！")
            # --- 修改後的統計顯示 ---
            print(f"📊 總計交回回饋的人數：{len(sorted_staff_list)} 位")
            print(f"👥 交件同仁清單：")
            # 每 10 個人換一行列印，避免名字太長擠在一起
            for i in range(0, len(sorted_staff_list), 10):
                print(f"   {', '.join(sorted_staff_list[i:i + 10])}")

            print(f"\n📝 總表狀況：保留完整名單（共 {len(final_df)} 列），並已填入測試資訊。")
            print("=" * 50)
        except PermissionError:
            print("\n❌ 錯誤：檔案被開啟中，請關閉後再執行。")
    else:
        print("⚠️ 找不到任何資料。")